import openpyxl
import json

path = r"src\entities\rooms\Highrise Cheatsheet.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

data = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    data[sheet_name] = rows

print(json.dumps(data, indent=2))
